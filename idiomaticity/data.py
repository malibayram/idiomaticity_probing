"""NCIMP dataset loading and the in-memory data model.

The paper's NCIMP dataset stores, for each noun compound (NC) and each example sentence, the
original sentence plus several *pre-substituted* sentences (one per probe), each with a ``_tag``
column marking the target token positions. We expose:

  * :class:`Substitution`   - one sentence + the (start, end) word span of the target tokens
  * :class:`NCInstance`     - one NC in one sentence: original + every probe's substitution(s)
  * :func:`load_dataset`    - load a directory of NCIMP/canonical data into ``list[NCInstance]``

Two on-disk formats are supported:
  1. **canonical** JSON (``*.json``) - the clean format this project emits/consumes; see
     ``data/sample/``. Recommended.
  2. **NCIMP CSV** - the original repo layout
     (``dataset/<LANG>/naturalistics_examplesent{1,2,3}.csv``, ``neutral.csv``), via
     :func:`load_ncimp_csv`.

Compositionality classes: ``I`` (idiomatic), ``PC`` (partly compositional), ``C``
(compositional). Context types: ``naturalistic`` (S1–S3) and ``neutral``.
"""

from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from .probes import Probe

Span = Tuple[int, int]

COMP_CLASSES = ("I", "PC", "C")
CONTEXT_TYPES = ("naturalistic", "neutral")


# ----------------------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------------------
@dataclass(frozen=True)
class Substitution:
    """A sentence with the target tokens (NC or a probe replacement) located by word span."""

    sentence: str
    span: Span

    def target_text(self) -> str:
        return " ".join(self.sentence.split()[self.span[0] : self.span[1]])


@dataclass
class NCInstance:
    """One noun compound observed in one sentence, with all probe substitutions."""

    nc: str
    lang: str
    context: str  # "naturalistic" | "neutral"
    comp_class: str  # "I" | "PC" | "C"
    comp_score: Optional[float]  # human compositionality, 0=idiomatic .. 5=compositional
    original: Substitution
    probes: Dict[Probe, List[Substitution]] = field(default_factory=dict)
    sent_id: Optional[str] = None  # e.g. "S1"

    def probe_subs(self, probe: Probe) -> List[Substitution]:
        return self.probes.get(probe, [])


# ----------------------------------------------------------------------------------------
# Tag / span helpers
# ----------------------------------------------------------------------------------------
def span_from_tag(sentence: str, tag: str | Sequence) -> Optional[Span]:
    """Return the (start, end) word span marked truthy in a ``tag``.

    Accepts the NCIMP list-literal format (``"[False, True, True, False]"``), a space-separated
    string (``"0 0 1 1 0"``), BIO-ish tags (``"O O B I O"``), or an already-parsed list.
    Returns the first contiguous marked run as a half-open word span, or ``None``.
    """
    if tag is None:
        return None
    items = _parse_tag(tag)
    if not items:
        return None
    marked = [i for i, t in enumerate(items) if _is_marked(t)]
    if not marked:
        return None
    start = marked[0]
    end = start
    for i in marked:
        if i == end or i == end + 1:
            end = i
        else:
            break
    return (start, end + 1)


def _parse_tag(tag: str | Sequence) -> list:
    """Normalize a tag value into a list of per-word markers."""
    if not isinstance(tag, str):
        return list(tag)
    s = tag.strip()
    if s.startswith("[") and s.endswith("]"):
        try:
            import ast

            return list(ast.literal_eval(s))
        except (ValueError, SyntaxError):
            # Fallback: strip brackets and split on commas.
            return [t.strip() for t in s[1:-1].split(",")]
    return s.split()


def _is_marked(tok) -> bool:
    if isinstance(tok, bool):
        return tok
    s = str(tok).strip().upper()
    return s not in ("", "0", "O", "FALSE", "NONE", "_")


def locate_phrase(sentence: str, phrase: str) -> Optional[Span]:
    """Find ``phrase`` as a contiguous run of words in ``sentence``; return its word span."""
    if not phrase:
        return None
    words = sentence.split()
    target = phrase.split()
    n = len(target)
    low_words = [w.lower().strip(".,;:!?") for w in words]
    low_target = [w.lower().strip(".,;:!?") for w in target]
    for i in range(len(words) - n + 1):
        if low_words[i : i + n] == low_target:
            return (i, i + n)
    return None


def make_substitution(sentence: str, tag=None, phrase: str = "") -> Optional[Substitution]:
    """Build a Substitution, preferring the explicit tag, falling back to phrase search."""
    sentence = (sentence or "").strip()
    if not sentence:
        return None
    span = span_from_tag(sentence, tag) if tag is not None else None
    if span is None and phrase:
        span = locate_phrase(sentence, phrase)
    if span is None:
        # last resort: whole sentence as the span
        span = (0, len(sentence.split()))
    return Substitution(sentence=sentence, span=span)


# ----------------------------------------------------------------------------------------
# Canonical JSON loader
# ----------------------------------------------------------------------------------------
def load_canonical_json(path: str) -> List[NCInstance]:
    with open(path, "r", encoding="utf-8") as fh:
        raw = json.load(fh)
    items: List[NCInstance] = []
    for rec in raw:
        original = Substitution(rec["original"]["sentence"], tuple(rec["original"]["span"]))
        probes: Dict[Probe, List[Substitution]] = {}
        for pname, subs in rec.get("probes", {}).items():
            probe = Probe(pname)
            probes[probe] = [Substitution(s["sentence"], tuple(s["span"])) for s in subs]
        items.append(
            NCInstance(
                nc=rec["nc"],
                lang=rec["lang"],
                context=rec["context"],
                comp_class=rec["comp_class"],
                comp_score=rec.get("comp_score"),
                original=original,
                probes=probes,
                sent_id=rec.get("sent_id"),
            )
        )
    return items


def dump_canonical_json(items: Iterable[NCInstance], path: str) -> None:
    out = []
    for it in items:
        out.append(
            {
                "nc": it.nc,
                "lang": it.lang,
                "context": it.context,
                "comp_class": it.comp_class,
                "comp_score": it.comp_score,
                "sent_id": it.sent_id,
                "original": {"sentence": it.original.sentence, "span": list(it.original.span)},
                "probes": {
                    p.value: [{"sentence": s.sentence, "span": list(s.span)} for s in subs]
                    for p, subs in it.probes.items()
                },
            }
        )
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)


# ----------------------------------------------------------------------------------------
# NCIMP CSV adapter (original repo format)
# ----------------------------------------------------------------------------------------
# Column names in the original repo (see repo `utils.py`):
_COL = {
    "orig_nat": "original sentence",
    "orig_neut": "neutral sentence",
    "syn": "synonym for compound",
    "head": "original head only",
    "mod": "original modifier only",
    "words_syn": "synonym both",
    "rand": "nc rand freq sentence",  # + 1..5
}


def _get(row: dict, key: str, default: str = "") -> str:
    return (row.get(key) or default).strip()


def load_ncimp_csv(
    csv_path: str,
    lang: str,
    context: str,
    comp_scores: Optional[Dict[str, float]] = None,
    comp_classes: Optional[Dict[str, str]] = None,
    sent_id: Optional[str] = None,
) -> List[NCInstance]:
    """Load one NCIMP CSV file into NCInstances.

    ``comp_scores`` / ``comp_classes`` map an NC string to its human Comp score / class
    (loaded separately from the xlsx - see :func:`load_comp_scores`).
    """
    comp_scores = comp_scores or {}
    comp_classes = comp_classes or {}
    orig_col = _COL["orig_neut"] if context == "neutral" else _COL["orig_nat"]
    items: List[NCInstance] = []

    # utf-8-sig strips the BOM on the first column name ("﻿compound").
    with open(csv_path, "r", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        # Surface form (as it appears in the sentence, may be plural).
        nc_col = "compound noun" if "compound noun" in reader.fieldnames else reader.fieldnames[0]
        # Canonical form (singular lemma) used to join with the xlsx scores.
        canon_col = "compound" if "compound" in reader.fieldnames else nc_col
        for row in reader:
            nc = _get(row, nc_col)
            canon = _get(row, canon_col) or nc
            orig_sentence = _get(row, orig_col)
            if not orig_sentence:
                continue
            # The original-position tag column is named "original sentence_tag" in both the
            # naturalistic and neutral files.
            original = make_substitution(
                orig_sentence, row.get("original sentence_tag"), phrase=nc
            )
            if original is None:
                continue

            probes: Dict[Probe, List[Substitution]] = {}

            syn = make_substitution(
                _get(row, _COL["syn"]), row.get(_COL["syn"] + "_tag")
            )
            if syn:
                probes[Probe.SYN] = [syn]

            # P_Comp: prefer the head-only sentence (semantic head of the NC).
            comp_sub = make_substitution(
                _get(row, _COL["head"]), row.get(_COL["head"] + "_tag")
            ) or make_substitution(_get(row, _COL["mod"]), row.get(_COL["mod"] + "_tag"))
            if comp_sub:
                probes[Probe.COMP] = [comp_sub]

            words_syn = make_substitution(
                _get(row, _COL["words_syn"]), row.get(_COL["words_syn"] + "_tag")
            )
            if words_syn:
                probes[Probe.WORDS_SYN] = [words_syn]

            rand_subs = []
            for k in range(1, 6):
                s = make_substitution(
                    _get(row, f"{_COL['rand']}{k}"), row.get(f"{_COL['rand']}{k}_tag")
                )
                if s:
                    rand_subs.append(s)
            if rand_subs:
                probes[Probe.RAND] = rand_subs

            key = canon.lower()
            items.append(
                NCInstance(
                    nc=nc,
                    lang=lang,
                    context=context,
                    comp_class=comp_classes.get(key, "?"),
                    comp_score=comp_scores.get(key),
                    original=original,
                    probes=probes,
                    sent_id=sent_id,
                )
            )
    return items


def load_comp_scores(
    xlsx_path: str, lang: Optional[str] = None
) -> Tuple[Dict[str, float], Dict[str, str]]:
    """Load human compositionality scores + classes from the NCIMP xlsx.

    Real schema columns: ``language, experiment_type, file, compound, ClassType, ClassToken,
    CompositionalityType, CompositionalityToken``. We use the **type-level** score
    (``CompositionalityType``, 0=idiomatic..5=compositional) and class (``ClassType``), keyed
    by compound. ``ClassType='NC'`` means non-compositional → mapped to the paper's ``I``.

    ``lang`` filters by the ``language`` column ("en"/"pt"); ``None`` keeps all.
    Returns ``(scores, classes)`` keyed by NC string.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}, {}
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    lang_i = idx.get("language")
    nc_i = idx.get("compound")
    score_i = idx.get("CompositionalityType")
    class_i = idx.get("ClassType")
    if nc_i is None or score_i is None:
        return {}, {}

    target = lang.lower() if lang else None
    scores: Dict[str, float] = {}
    classes: Dict[str, str] = {}
    for row in rows:
        if nc_i >= len(row) or row[nc_i] is None:
            continue
        if target is not None and lang_i is not None:
            if str(row[lang_i]).strip().lower() != target:
                continue
        nc = str(row[nc_i]).strip().lower()
        try:
            scores[nc] = float(row[score_i])
        except (TypeError, ValueError):
            pass
        if class_i is not None and class_i < len(row) and row[class_i] is not None:
            classes[nc] = _normalize_class(str(row[class_i]))
    return scores, classes


def _normalize_class(value: str) -> str:
    v = value.strip().lower()
    if v in ("nc", "i") or v.startswith("idio") or v.startswith("non"):
        return "I"  # NC = non-compositional = idiomatic
    if v in ("pc",) or v.startswith("part"):
        return "PC"
    if v in ("c",) or v.startswith("comp"):
        return "C"
    return value.strip()


# ----------------------------------------------------------------------------------------
# Unified loader
# ----------------------------------------------------------------------------------------
def load_dataset(
    data_dir: str,
    lang: str = "EN",
    contexts: Sequence[str] = ("naturalistic",),
) -> List[NCInstance]:
    """Load a dataset directory.

    Resolution order:
      1. ``<data_dir>/<lang>.json`` or ``<data_dir>/*.json`` (canonical format), or
      2. NCIMP CSV layout ``<data_dir>/<LANG>/naturalistics_examplesent{1,2,3}.csv`` and
         ``neutral.csv`` (+ optional ``human_compositionality scores.xlsx``).
    """
    # 1. Canonical JSON
    json_candidates = [
        os.path.join(data_dir, f"{lang}.json"),
        os.path.join(data_dir, f"{lang.lower()}.json"),
        os.path.join(data_dir, "dataset.json"),
    ]
    for jc in json_candidates:
        if os.path.isfile(jc):
            items = load_canonical_json(jc)
            return [it for it in items if it.lang == lang and it.context in contexts] or items

    # 2. NCIMP CSV
    lang_dir = os.path.join(data_dir, lang)
    if os.path.isdir(lang_dir):
        comp_scores, comp_classes = {}, {}
        for fname in os.listdir(data_dir):
            if fname.lower().endswith(".xlsx"):
                comp_scores, comp_classes = load_comp_scores(
                    os.path.join(data_dir, fname), lang=lang
                )
                break
        items: List[NCInstance] = []
        for context in contexts:
            if context == "neutral":
                files = [("neutral.csv", None)]
            else:
                files = [
                    (f"naturalistics_examplesent{i}.csv", f"S{i}") for i in (1, 2, 3)
                ]
            for fname, sid in files:
                fpath = os.path.join(lang_dir, fname)
                if os.path.isfile(fpath):
                    items.extend(
                        load_ncimp_csv(
                            fpath, lang, context, comp_scores, comp_classes, sid
                        )
                    )
        return items

    raise FileNotFoundError(
        f"No canonical JSON or NCIMP CSV layout found under {data_dir!r} for lang={lang!r}"
    )
